from datetime import datetime, timedelta, date as date_type
from decimal import Decimal, InvalidOperation
from io import BytesIO
import csv
import io
import requests as http_requests
import openpyxl
import xlrd
from openpyxl.styles import Font
from flask import (Blueprint, render_template, redirect, url_for, request, flash, g,
                    jsonify, abort, current_app)
from flask_login import login_required, current_user
from sqlalchemy import func, extract
from ..extensions import db
from ..models import (Business, User, Product, Category, StockEntry,
                      Sale, SaleItem, WorkSession, PriceHistory, Expense, Subscription)
from ..config import PLAN_LIMITS, Config
from ..utils import load_business_or_404, admin_required, get_date_range, sales_summary, slugify, send_excel_file

admin_bp = Blueprint('admin', __name__)


def _load_biz(slug):
    biz = load_business_or_404(slug)
    if not current_user.is_authenticated or current_user.business_id != biz.id or current_user.role != 'admin':
        abort(403)
    g.business = biz
    return biz


# ─── Dashboard ───────────────────────────────────────────────────────────────

@admin_bp.route('/<slug>/admin/')
@login_required
def dashboard(slug):
    biz = _load_biz(slug)
    today_start, today_end = get_date_range('today')
    week_start, week_end = get_date_range('week')
    month_start, month_end = get_date_range('month')

    today_summary = sales_summary(biz.id, today_start, today_end)
    week_summary = sales_summary(biz.id, week_start, week_end)
    month_summary = sales_summary(biz.id, month_start, month_end)

    total_products = Product.query.filter_by(business_id=biz.id, is_active=True).count()
    low_stock = Product.query.filter(
        Product.business_id == biz.id,
        Product.is_active == True,
        Product.quantity_in_stock <= Product.reorder_level
    ).count()
    active_sessions = WorkSession.query.filter_by(business_id=biz.id, is_active=True).count()

    stock_value = db.session.query(
        func.sum(Product.cost_price * Product.quantity_in_stock)
    ).filter_by(business_id=biz.id, is_active=True).scalar() or 0

    # last 7 days chart data
    chart_labels = []
    chart_data = []
    for i in range(6, -1, -1):
        d = (datetime.utcnow() - timedelta(days=i)).date()
        label = d.strftime('%a %d')
        start = datetime.combine(d, datetime.min.time())
        end = datetime.combine(d, datetime.max.time())
        rev = db.session.query(func.sum(Sale.total_amount)).filter(
            Sale.business_id == biz.id,
            Sale.created_at >= start,
            Sale.created_at <= end
        ).scalar() or 0
        chart_labels.append(label)
        chart_data.append(float(rev))

    recent_sales = Sale.query.filter_by(business_id=biz.id).order_by(Sale.created_at.desc()).limit(10).all()

    return render_template('admin/dashboard.html',
                           business=biz,
                           today=today_summary,
                           week=week_summary,
                           month=month_summary,
                           total_products=total_products,
                           low_stock=low_stock,
                           active_sessions=active_sessions,
                           stock_value=float(stock_value),
                           chart_labels=chart_labels,
                           chart_data=chart_data,
                           recent_sales=recent_sales)


# ─── Products ────────────────────────────────────────────────────────────────

def _apply_product_filters(query, cat_id, q, stock_filter, exp_filter):
    if cat_id:
        query = query.filter_by(category_id=cat_id)
    if q:
        query = query.filter(Product.name.ilike(f'%{q}%'))
    if stock_filter == 'low':
        query = query.filter(Product.quantity_in_stock <= Product.reorder_level)
    if exp_filter == 'expired':
        query = query.filter(Product.expiry_date.isnot(None), Product.expiry_date < date_type.today())
    elif exp_filter == 'expiring_soon':
        today = date_type.today()
        query = query.filter(Product.expiry_date.isnot(None),
                             Product.expiry_date >= today,
                             Product.expiry_date <= today + timedelta(days=30))
    return query


@admin_bp.route('/<slug>/admin/products')
@login_required
def products(slug):
    biz = _load_biz(slug)
    cats = Category.query.filter_by(business_id=biz.id).all()
    cat_id = request.args.get('cat', type=int)
    q = request.args.get('q', '').strip()
    stock_filter = request.args.get('stock', '').strip()
    exp_filter = request.args.get('exp', '').strip()
    query = _apply_product_filters(Product.query.filter_by(business_id=biz.id),
                                   cat_id, q, stock_filter, exp_filter)
    prods = query.order_by(Product.name).all()
    has_filter = bool(q or cat_id or stock_filter or exp_filter)
    return render_template('admin/products.html', business=biz, products=prods,
                           categories=cats, selected_cat=cat_id, q=q,
                           stock_filter=stock_filter, exp_filter=exp_filter,
                           has_filter=has_filter)


@admin_bp.route('/<slug>/admin/products/export')
@login_required
def export_products(slug):
    """Export the current (optionally filtered) product list to an .xlsx file."""
    biz = _load_biz(slug)
    cat_id = request.args.get('cat', type=int)
    q = request.args.get('q', '').strip()
    stock_filter = request.args.get('stock', '').strip()
    exp_filter = request.args.get('exp', '').strip()
    query = _apply_product_filters(Product.query.filter_by(business_id=biz.id),
                                   cat_id, q, stock_filter, exp_filter)
    prods = query.order_by(Product.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'
    headers = ['Name', 'Barcode', 'Category', 'Unit Price', 'Cost Price',
               'Quantity In Stock', 'Unit', 'Reorder Level',
               'Manufacture Date', 'Expiry Date', 'Description', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for p in prods:
        ws.append([
            p.name, p.barcode, p.category.name if p.category else '',
            float(p.unit_price), float(p.cost_price), p.quantity_in_stock,
            p.unit, p.reorder_level, p.manufacture_date, p.expiry_date,
            p.description or '', 'Active' if p.is_active else 'Inactive',
        ])
        r = ws.max_row
        ws.cell(row=r, column=9).number_format = 'DD/MM/YYYY'
        ws.cell(row=r, column=10).number_format = 'DD/MM/YYYY'

    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(12, length + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{biz.slug}-products-{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_excel_file(buf, filename)


@admin_bp.route('/<slug>/admin/products/check-barcode')
@login_required
def check_barcode(slug):
    """Live duplicate-barcode lookup used by the add/edit product form as you scan/type."""
    biz = _load_biz(slug)
    barcode = request.args.get('barcode', '').strip()
    exclude_id = request.args.get('exclude_id', type=int)

    if not barcode:
        return jsonify(exists=False)

    query = Product.query.filter_by(business_id=biz.id, barcode=barcode)
    if exclude_id:
        query = query.filter(Product.id != exclude_id)
    existing = query.first()

    return jsonify(exists=existing is not None, product_name=existing.name if existing else None)


@admin_bp.route('/<slug>/admin/products/new', methods=['GET', 'POST'])
@login_required
def new_product(slug):
    biz = _load_biz(slug)
    sub = biz.subscription
    features = sub.features if sub else PLAN_LIMITS['free']
    max_products = features.get('max_products')
    if max_products is not None:
        current_count = Product.query.filter_by(business_id=biz.id, is_active=True).count()
        if current_count >= max_products:
            flash(f'Your {features["label"]} plan allows up to {max_products} products. '
                  f'Upgrade to add more.', 'warning')
            return redirect(url_for('admin.upgrade', slug=slug))

    cats = Category.query.filter_by(business_id=biz.id).all()
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Product name is required.', 'danger')
            return render_template('admin/product_form.html', business=biz, categories=cats, product=None)

        barcode = request.form.get('barcode', '').strip()
        if not barcode:
            flash('Barcode is required.', 'danger')
            return render_template('admin/product_form.html', business=biz, categories=cats, product=None)
        if Product.query.filter_by(business_id=biz.id, barcode=barcode).first():
            flash(f'Barcode "{barcode}" is already assigned to another product.', 'danger')
            return render_template('admin/product_form.html', business=biz, categories=cats, product=None)

        mfg_raw = request.form.get('manufacture_date', '').strip()
        exp_raw = request.form.get('expiry_date', '').strip()
        try:
            mfg_date = datetime.strptime(mfg_raw, '%d/%m/%Y').date() if mfg_raw else None
            exp_date = datetime.strptime(exp_raw, '%d/%m/%Y').date() if exp_raw else None
        except ValueError:
            flash('Invalid date format. Please use dd/mm/yyyy.', 'danger')
            return render_template('admin/product_form.html', business=biz, categories=cats, product=None)

        prod = Product(
            business_id=biz.id,
            name=name,
            barcode=barcode,
            manufacture_date=mfg_date,
            expiry_date=exp_date,
            description=request.form.get('description', '').strip(),
            unit_price=Decimal(request.form.get('unit_price', '0') or '0'),
            cost_price=Decimal(request.form.get('cost_price', '0') or '0'),
            reorder_level=int(request.form.get('reorder_level', 5) or 5),
            unit=request.form.get('unit', 'piece').strip(),
            category_id=request.form.get('category_id', type=int) or None,
        )
        db.session.add(prod)
        db.session.commit()
        flash(f'Product "{name}" created.', 'success')
        return redirect(url_for('admin.products', slug=slug))
    return render_template('admin/product_form.html', business=biz, categories=cats, product=None)


@admin_bp.route('/<slug>/admin/products/<int:product_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_product(slug, product_id):
    biz = _load_biz(slug)
    prod = Product.query.filter_by(id=product_id, business_id=biz.id).first_or_404()
    cats = Category.query.filter_by(business_id=biz.id).all()

    if request.method == 'POST':
        old_price = prod.unit_price
        prod.name = request.form.get('name', prod.name).strip()
        new_barcode = request.form.get('barcode', '').strip()
        if not new_barcode:
            flash('Barcode is required.', 'danger')
            return render_template('admin/product_form.html', business=biz, categories=cats, product=prod)
        if new_barcode != prod.barcode:
            conflict = Product.query.filter_by(business_id=biz.id, barcode=new_barcode).first()
            if conflict:
                flash(f'Barcode "{new_barcode}" is already assigned to "{conflict.name}".', 'danger')
                return render_template('admin/product_form.html', business=biz, categories=cats, product=prod)
        prod.barcode = new_barcode
        prod.description = request.form.get('description', '').strip()
        prod.unit_price = Decimal(request.form.get('unit_price', str(prod.unit_price)) or '0')
        prod.cost_price = Decimal(request.form.get('cost_price', str(prod.cost_price)) or '0')
        prod.reorder_level = int(request.form.get('reorder_level', prod.reorder_level) or prod.reorder_level)
        prod.unit = request.form.get('unit', prod.unit).strip()
        prod.category_id = request.form.get('category_id', type=int) or None
        prod.is_active = bool(request.form.get('is_active'))
        mfg_raw = request.form.get('manufacture_date', '').strip()
        exp_raw = request.form.get('expiry_date', '').strip()
        try:
            prod.manufacture_date = datetime.strptime(mfg_raw, '%d/%m/%Y').date() if mfg_raw else None
            prod.expiry_date = datetime.strptime(exp_raw, '%d/%m/%Y').date() if exp_raw else None
        except ValueError:
            flash('Invalid date format. Please use dd/mm/yyyy.', 'danger')
            return render_template('admin/product_form.html', business=biz, categories=cats, product=prod)
        prod.updated_at = datetime.utcnow()

        if prod.unit_price != old_price:
            ph = PriceHistory(
                product_id=prod.id,
                old_price=old_price,
                new_price=prod.unit_price,
                changed_by=current_user.id,
            )
            db.session.add(ph)

        db.session.commit()
        flash(f'Product "{prod.name}" updated.', 'success')
        return redirect(url_for('admin.products', slug=slug))

    return render_template('admin/product_form.html', business=biz, categories=cats, product=prod)


@admin_bp.route('/<slug>/admin/products/<int:product_id>/toggle', methods=['POST'])
@login_required
def toggle_product(slug, product_id):
    biz = _load_biz(slug)
    prod = Product.query.filter_by(id=product_id, business_id=biz.id).first_or_404()
    prod.is_active = not prod.is_active
    db.session.commit()
    status = 'activated' if prod.is_active else 'deactivated'
    flash(f'Product "{prod.name}" {status}.', 'success')
    return redirect(url_for('admin.products', slug=slug))


BULK_UPLOAD_COLUMNS = [
    'name', 'barcode', 'category', 'unit_price', 'cost_price',
    'quantity_in_stock', 'unit', 'reorder_level',
    'manufacture_date', 'expiry_date', 'description',
]

# Maps common header variants (lowercased, spaces -> underscores) seen in real-world stock
# lists onto the canonical column names above, so users don't have to rename their columns.
BULK_UPLOAD_HEADER_ALIASES = {
    'product_name': 'name', 'product': 'name', 'item_name': 'name', 'item': 'name',
    'sku': 'barcode',
    'selling_price': 'unit_price', 'price': 'unit_price', 'sale_price': 'unit_price',
    'buying_price': 'cost_price', 'purchase_price': 'cost_price',
    'qty': 'quantity_in_stock', 'quantity': 'quantity_in_stock',
    'stock': 'quantity_in_stock', 'stock_qty': 'quantity_in_stock',
    'reorder': 'reorder_level', 'reorder_point': 'reorder_level',
    'mfg_date': 'manufacture_date', 'expiry': 'expiry_date', 'exp_date': 'expiry_date',
}


def _cell_text(value):
    """Stringify an openpyxl cell value, collapsing whole-number floats (e.g. barcodes)."""
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_date(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date_type):
        return value
    try:
        return datetime.strptime(_cell_text(value), '%Y-%m-%d').date()
    except ValueError:
        return None


BULK_UPLOAD_EXTENSIONS = ('.xlsx', '.xlsm', '.xls', '.csv')


def _read_bulk_upload_rows(file):
    """Parse an uploaded bulk-upload file (.xlsx/.xlsm/.xls/.csv) into a list of row tuples.

    Raises ValueError with a user-facing message if the file can't be read.
    """
    filename = file.filename.lower()

    if filename.endswith('.csv'):
        file.stream.seek(0)
        raw = file.stream.read()
        for encoding in ('utf-8-sig', 'latin-1'):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError('Could not decode the CSV file. Please save it as UTF-8 CSV and try again.')
        try:
            return [tuple(row) for row in csv.reader(io.StringIO(text))]
        except csv.Error as e:
            raise ValueError(f'Could not read the CSV file ({e}).')

    if filename.endswith('.xls'):
        file.stream.seek(0)
        try:
            book = xlrd.open_workbook(file_contents=file.stream.read())
            sheet = book.sheet_by_index(0)
            return [tuple(sheet.row_values(r)) for r in range(sheet.nrows)]
        except Exception as e:
            raise ValueError(f'Could not read the .xls file ({e}). Please make sure it is a valid Excel 97-2003 file.')

    # .xlsx / .xlsm
    # Read the whole upload into a real in-memory buffer first. Passing file.stream
    # straight to openpyxl breaks on some server Python builds where the underlying
    # SpooledTemporaryFile doesn't implement .seekable() (needed pre-3.11) — zipfile
    # (which .xlsx parsing relies on) calls that method and raises AttributeError.
    file.stream.seek(0)
    raw = file.stream.read()

    last_err = None
    for read_only in (True, False):
        try:
            wb = openpyxl.load_workbook(BytesIO(raw), data_only=True, read_only=read_only)
            ws = wb.active
            if ws is None:
                last_err = 'the workbook has no readable worksheet'
                continue
            return list(ws.iter_rows(values_only=True))
        except Exception as e:
            # read_only mode is strict about worksheet metadata (e.g. a missing/incorrect
            # <dimension> tag) that some non-Excel tools (Google Sheets, LibreOffice, WPS)
            # omit. Fall back to normal load, which is more tolerant.
            last_err = e
            continue

    raise ValueError(f'Could not read the file ({last_err}). The file may be corrupted — try opening it in '
                      f'Excel or Google Sheets and saving it as a new .xlsx file.')


@admin_bp.route('/<slug>/admin/products/bulk-upload', methods=['GET', 'POST'])
@login_required
def bulk_upload_products(slug):
    biz = _load_biz(slug)
    sub = biz.subscription
    features = sub.features if sub else PLAN_LIMITS['free']
    max_products = features.get('max_products')

    if request.method == 'POST':
        file = request.files.get('excel_file')
        if not file or not file.filename:
            flash('Please choose a file to upload.', 'danger')
            return render_template('admin/product_bulk_upload.html', business=biz, max_products=max_products)
        if not file.filename.lower().endswith(BULK_UPLOAD_EXTENSIONS):
            flash('Supported formats: .xlsx, .xlsm, .xls, .csv.', 'danger')
            return render_template('admin/product_bulk_upload.html', business=biz, max_products=max_products)

        try:
            rows = _read_bulk_upload_rows(file)
        except ValueError as e:
            flash(str(e), 'danger')
            return render_template('admin/product_bulk_upload.html', business=biz, max_products=max_products)

        if not rows:
            flash('The uploaded file is empty.', 'danger')
            return render_template('admin/product_bulk_upload.html', business=biz, max_products=max_products)

        header_raw = [_cell_text(c).lower().replace(' ', '_') for c in rows[0]]
        header = [BULK_UPLOAD_HEADER_ALIASES.get(h, h) for h in header_raw]
        if 'name' not in header or 'barcode' not in header:
            flash('The file must contain at least "name" (or "Product Name") and "barcode" columns.', 'danger')
            return render_template('admin/product_bulk_upload.html', business=biz, max_products=max_products)
        if 'unit_price' not in header:
            flash('The file must contain a selling price column ("unit_price" or "Selling Price").', 'danger')
            return render_template('admin/product_bulk_upload.html', business=biz, max_products=max_products)
        col_idx = {col: i for i, col in enumerate(header) if col}

        def cell(row, key):
            idx = col_idx.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        existing_count = Product.query.filter_by(business_id=biz.id, is_active=True).count()
        existing_barcodes = {b for (b,) in db.session.query(Product.barcode).filter_by(business_id=biz.id).all()}
        cats_by_name = {c.name.lower(): c for c in Category.query.filter_by(business_id=biz.id).all()}

        created = 0
        skipped = []
        seen_in_file = set()
        limit_hit = False

        for i, row in enumerate(rows[1:], start=2):
            if row is None or all(v is None or _cell_text(v) == '' for v in row):
                continue

            if limit_hit:
                skipped.append(f'Row {i}: skipped — plan limit of {max_products} products reached.')
                continue

            name = _cell_text(cell(row, 'name'))
            barcode = _cell_text(cell(row, 'barcode'))
            unit_price_raw = _cell_text(cell(row, 'unit_price'))
            if not name or not barcode:
                skipped.append(f'Row {i}: missing name or barcode.')
                continue
            if not unit_price_raw:
                skipped.append(f'Row {i}: missing selling price.')
                continue
            if barcode in existing_barcodes or barcode in seen_in_file:
                skipped.append(f'Row {i}: barcode "{barcode}" already exists.')
                continue
            if max_products is not None and existing_count + created >= max_products:
                skipped.append(f'Row {i}: skipped — plan limit of {max_products} products reached.')
                limit_hit = True
                continue

            try:
                unit_price = Decimal(unit_price_raw)
                cost_price = Decimal(_cell_text(cell(row, 'cost_price')) or '0')
            except InvalidOperation:
                skipped.append(f'Row {i}: invalid selling price or cost price.')
                continue

            try:
                qty = int(float(_cell_text(cell(row, 'quantity_in_stock')) or 0))
                reorder_level = int(float(_cell_text(cell(row, 'reorder_level')) or 5))
            except ValueError:
                skipped.append(f'Row {i}: invalid quantity or reorder level.')
                continue

            mfg_raw = cell(row, 'manufacture_date')
            exp_raw = cell(row, 'expiry_date')
            if (mfg_raw not in (None, '') and _cell_date(mfg_raw) is None) or \
               (exp_raw not in (None, '') and _cell_date(exp_raw) is None):
                skipped.append(f'Row {i}: invalid date format (use YYYY-MM-DD).')
                continue

            cat_name = _cell_text(cell(row, 'category'))
            category_id = None
            if cat_name:
                cat = cats_by_name.get(cat_name.lower())
                if not cat:
                    cat = Category(business_id=biz.id, name=cat_name)
                    db.session.add(cat)
                    db.session.flush()
                    cats_by_name[cat_name.lower()] = cat
                category_id = cat.id

            unit = _cell_text(cell(row, 'unit')) or 'piece'

            prod = Product(
                business_id=biz.id,
                name=name,
                barcode=barcode,
                description=_cell_text(cell(row, 'description')),
                unit_price=unit_price,
                cost_price=cost_price,
                quantity_in_stock=qty,
                reorder_level=reorder_level,
                unit=unit,
                category_id=category_id,
                manufacture_date=_cell_date(mfg_raw),
                expiry_date=_cell_date(exp_raw),
            )
            db.session.add(prod)
            seen_in_file.add(barcode)
            created += 1

        db.session.commit()

        if created and not skipped:
            flash(f'{created} product(s) imported successfully.', 'success')
            return redirect(url_for('admin.products', slug=slug))

        if created and skipped:
            flash(f'{created} product(s) imported. {len(skipped)} row(s) skipped — see details below.', 'warning')
        elif not created:
            flash(f'No products were imported. {len(skipped)} row(s) skipped — see details below.', 'danger')

        return render_template('admin/product_bulk_upload.html', business=biz, max_products=max_products,
                               created=created, skipped=skipped)

    return render_template('admin/product_bulk_upload.html', business=biz, max_products=max_products)


@admin_bp.route('/<slug>/admin/products/bulk-upload/template')
@login_required
def bulk_upload_template(slug):
    biz = _load_biz(slug)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'
    ws.append(BULK_UPLOAD_COLUMNS)
    ws.append(['Coca-Cola 50cl', '5449000000996', 'Beverages', 500, 350,
               100, 'bottle', 10, '2026-01-15', '2026-12-31', 'Soft drink'])
    for col_cells in ws.columns:
        length = max(len(_cell_text(c.value)) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(12, length + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_excel_file(buf, 'product_bulk_upload_template.xlsx')


# ─── Categories ──────────────────────────────────────────────────────────────

@admin_bp.route('/<slug>/admin/categories/new', methods=['POST'])
@login_required
def new_category(slug):
    biz = _load_biz(slug)
    name = request.form.get('name', '').strip()
    if name:
        cat = Category(business_id=biz.id, name=name)
        db.session.add(cat)
        db.session.commit()
        flash(f'Category "{name}" added.', 'success')
    return redirect(request.referrer or url_for('admin.products', slug=slug))


# ─── Stock ───────────────────────────────────────────────────────────────────

@admin_bp.route('/<slug>/admin/stock')
@login_required
def stock(slug):
    biz = _load_biz(slug)
    entries = (StockEntry.query
               .filter_by(business_id=biz.id)
               .join(Product)
               .order_by(StockEntry.created_at.desc())
               .limit(200).all())
    all_products = Product.query.filter_by(business_id=biz.id, is_active=True).order_by(Product.name).all()
    low_stock_items = [p for p in all_products if p.is_low_stock]
    products_stock_json = [{'name': p.name, 'qty': p.quantity_in_stock} for p in all_products]
    return render_template('admin/stock.html', business=biz, entries=entries,
                           all_products=all_products, low_stock_items=low_stock_items,
                           products_stock_json=products_stock_json)


@admin_bp.route('/<slug>/admin/stock/bulk-restock', methods=['POST'])
@login_required
def bulk_restock(slug):
    """Set every active product at or below a chosen stock threshold to the same quantity."""
    biz = _load_biz(slug)
    threshold = request.form.get('threshold', type=int)
    target_qty = request.form.get('target_qty', type=int)
    reason = request.form.get('reason', '').strip()

    if threshold is None or threshold < 0:
        flash('Enter a valid stock threshold (0 or more).', 'danger')
        return redirect(url_for('admin.stock', slug=slug))
    if target_qty is None or target_qty <= 0:
        flash('Enter a valid quantity greater than 0.', 'danger')
        return redirect(url_for('admin.stock', slug=slug))

    matching_products = Product.query.filter(
        Product.business_id == biz.id,
        Product.is_active == True,
        Product.quantity_in_stock <= threshold
    ).all()

    if not matching_products:
        flash(f'No products currently have stock at or below {threshold}.', 'info')
        return redirect(url_for('admin.stock', slug=slug))

    notes = f'[BULK RESTOCK] {reason}' if reason else f'[BULK RESTOCK] Items at or below {threshold} units set to {target_qty}'
    for prod in matching_products:
        delta = target_qty - prod.quantity_in_stock
        prod.quantity_in_stock = target_qty
        if delta != 0:
            db.session.add(StockEntry(
                business_id=biz.id,
                product_id=prod.id,
                user_id=current_user.id,
                quantity=delta,
                unit_cost=prod.cost_price,
                total_cost=prod.cost_price * delta,
                notes=notes,
            ))

    db.session.commit()
    flash(f'{len(matching_products)} product(s) at or below {threshold} units updated to {target_qty}.', 'success')
    return redirect(url_for('admin.stock', slug=slug))


@admin_bp.route('/<slug>/admin/stock/adjust', methods=['POST'])
@login_required
def adjust_stock(slug):
    """Admin-only direct stock adjustment (not an immutable stock entry)."""
    biz = _load_biz(slug)
    product_id = request.form.get('product_id', type=int)
    quantity = request.form.get('quantity', type=int)
    reason = request.form.get('reason', '').strip()

    prod = Product.query.filter_by(id=product_id, business_id=biz.id).first_or_404()
    if quantity is None:
        flash('Quantity is required.', 'danger')
        return redirect(url_for('admin.stock', slug=slug))

    prod.quantity_in_stock += quantity
    if prod.quantity_in_stock < 0:
        prod.quantity_in_stock = 0

    entry = StockEntry(
        business_id=biz.id,
        product_id=prod.id,
        user_id=current_user.id,
        quantity=quantity,
        notes=f'[ADMIN ADJUSTMENT] {reason}',
    )
    db.session.add(entry)
    db.session.commit()
    flash(f'Stock adjusted for "{prod.name}".', 'success')
    return redirect(url_for('admin.stock', slug=slug))


# ─── Sales History ───────────────────────────────────────────────────────────

@admin_bp.route('/<slug>/admin/sales')
@login_required
def sales_history(slug):
    biz = _load_biz(slug)
    period = request.args.get('period', 'today')
    start, end = get_date_range(period)
    user_id = request.args.get('user_id', type=int)
    payment = request.args.get('payment', '').strip()

    query = Sale.query.filter(
        Sale.business_id == biz.id,
        Sale.created_at >= start,
        Sale.created_at <= end
    )
    if user_id:
        query = query.filter_by(user_id=user_id)
    if payment:
        query = query.filter_by(payment_method=payment)

    sale_list = query.order_by(Sale.created_at.desc()).all()
    summary = sales_summary(biz.id, start, end)
    sellers = User.query.filter_by(business_id=biz.id, is_active=True).all()

    return render_template('admin/sales.html', business=biz, sales=sale_list,
                           summary=summary, period=period, sellers=sellers,
                           selected_user=user_id, selected_payment=payment,
                           start=start, end=end)


@admin_bp.route('/<slug>/admin/sales/<int:sale_id>')
@login_required
def sale_detail(slug, sale_id):
    biz = _load_biz(slug)
    sale = Sale.query.filter_by(id=sale_id, business_id=biz.id).first_or_404()
    return render_template('admin/sale_detail.html', business=biz, sale=sale)


# ─── Reports / Analytics ─────────────────────────────────────────────────────

@admin_bp.route('/<slug>/admin/reports')
@login_required
def reports(slug):
    biz = _load_biz(slug)
    sub = biz.subscription
    features = sub.features if sub else PLAN_LIMITS['free']
    if not features.get('has_full_reports'):
        flash('Full financial reports are not available on your current plan. Upgrade to unlock them.', 'warning')
        return redirect(url_for('admin.upgrade', slug=slug))
    period = request.args.get('period', 'month')
    start, end = get_date_range(period)

    summary = sales_summary(biz.id, start, end)

    # ── COGS: qty_sold × current cost_price per product ──────────────────────
    cogs_raw = db.session.query(
        func.sum(SaleItem.quantity * Product.cost_price)
    ).join(Product, SaleItem.product_id == Product.id
    ).join(Sale, SaleItem.sale_id == Sale.id
    ).filter(
        Sale.business_id == biz.id,
        Sale.created_at >= start,
        Sale.created_at <= end
    ).scalar() or 0

    # ── Expenses for selected period ──────────────────────────────────────────
    period_expenses = (Expense.query
                       .filter(Expense.business_id == biz.id,
                               Expense.expense_date >= start.date(),
                               Expense.expense_date <= end.date())
                       .order_by(Expense.expense_date.desc()).all())

    total_expenses_raw = sum((e.amount for e in period_expenses), Decimal('0'))

    expense_by_cat = {}
    for e in period_expenses:
        expense_by_cat[e.category] = expense_by_cat.get(e.category, Decimal('0')) + e.amount

    # ── P&L ──────────────────────────────────────────────────────────────────
    revenue       = Decimal(str(summary['total_revenue']))
    cogs          = Decimal(str(cogs_raw))
    gross_profit  = revenue - cogs
    total_expenses = total_expenses_raw
    net_profit    = gross_profit - total_expenses
    gross_margin  = float(gross_profit / revenue * 100) if revenue else 0.0
    net_margin    = float(net_profit   / revenue * 100) if revenue else 0.0

    # ── Top products ──────────────────────────────────────────────────────────
    top_products = (db.session.query(
        SaleItem.product_name,
        func.sum(SaleItem.subtotal).label('revenue'),
        func.sum(SaleItem.quantity).label('qty_sold')
    ).join(Sale)
    .filter(Sale.business_id == biz.id, Sale.created_at >= start, Sale.created_at <= end)
    .group_by(SaleItem.product_name)
    .order_by(func.sum(SaleItem.subtotal).desc())
    .limit(10).all())

    # ── Sales by rep ──────────────────────────────────────────────────────────
    sales_by_rep = (db.session.query(
        User.first_name, User.last_name, User.username,
        func.count(Sale.id).label('count'),
        func.sum(Sale.total_amount).label('revenue')
    ).join(Sale, Sale.user_id == User.id)
    .filter(Sale.business_id == biz.id, Sale.created_at >= start, Sale.created_at <= end)
    .group_by(User.id)
    .order_by(func.sum(Sale.total_amount).desc()).all())

    # ── Payment breakdown ─────────────────────────────────────────────────────
    payment_breakdown = (db.session.query(
        Sale.payment_method,
        func.count(Sale.id).label('count'),
        func.sum(Sale.total_amount).label('revenue')
    ).filter(Sale.business_id == biz.id, Sale.created_at >= start, Sale.created_at <= end)
    .group_by(Sale.payment_method).all())

    # ── Revenue trend chart (for Sales Analytics tab) ─────────────────────────
    chart_days = 7 if period in ('today', 'week') else (30 if period == 'month' else 12)
    chart_labels, chart_revenue = [], []
    if period == 'year':
        for m in range(1, 13):
            ms = datetime(datetime.utcnow().year, m, 1)
            me = (datetime(datetime.utcnow().year, m + 1, 1) if m < 12
                  else datetime(datetime.utcnow().year + 1, 1, 1)) - timedelta(seconds=1)
            rev = db.session.query(func.sum(Sale.total_amount)).filter(
                Sale.business_id == biz.id, Sale.created_at >= ms, Sale.created_at <= me
            ).scalar() or 0
            chart_labels.append(ms.strftime('%b'))
            chart_revenue.append(float(rev))
    else:
        for i in range(chart_days - 1, -1, -1):
            d = (datetime.utcnow() - timedelta(days=i)).date()
            ds = datetime.combine(d, datetime.min.time())
            de = datetime.combine(d, datetime.max.time())
            rev = db.session.query(func.sum(Sale.total_amount)).filter(
                Sale.business_id == biz.id, Sale.created_at >= ds, Sale.created_at <= de
            ).scalar() or 0
            chart_labels.append(d.strftime('%m/%d'))
            chart_revenue.append(float(rev))

    # ── Yearly P&L analysis (always current calendar year) ───────────────────
    current_year = datetime.utcnow().year
    yearly_labels, yearly_rev, yearly_cogs_l = [], [], []
    yearly_exp_l, yearly_gross_l, yearly_net_l = [], [], []

    for m in range(1, 13):
        ms = datetime(current_year, m, 1)
        me = (datetime(current_year, m + 1, 1) if m < 12
              else datetime(current_year + 1, 1, 1)) - timedelta(seconds=1)

        m_rev = float(db.session.query(func.sum(Sale.total_amount)).filter(
            Sale.business_id == biz.id, Sale.created_at >= ms, Sale.created_at <= me
        ).scalar() or 0)

        m_cogs = float(db.session.query(
            func.sum(SaleItem.quantity * Product.cost_price)
        ).join(Product, SaleItem.product_id == Product.id
        ).join(Sale, SaleItem.sale_id == Sale.id
        ).filter(Sale.business_id == biz.id, Sale.created_at >= ms, Sale.created_at <= me
        ).scalar() or 0)

        m_exp = float(db.session.query(func.sum(Expense.amount)).filter(
            Expense.business_id == biz.id,
            Expense.expense_date >= ms.date(),
            Expense.expense_date <= me.date()
        ).scalar() or 0)

        yearly_labels.append(ms.strftime('%b'))
        yearly_rev.append(m_rev)
        yearly_cogs_l.append(m_cogs)
        yearly_exp_l.append(m_exp)
        yearly_gross_l.append(round(m_rev - m_cogs, 2))
        yearly_net_l.append(round(m_rev - m_cogs - m_exp, 2))

    return render_template('admin/reports.html',
        business=biz, period=period, start=start, end=end,
        summary=summary,
        cogs=float(cogs),
        gross_profit=float(gross_profit),
        total_expenses=float(total_expenses),
        net_profit=float(net_profit),
        gross_margin=gross_margin,
        net_margin=net_margin,
        expense_by_cat={k: float(v) for k, v in expense_by_cat.items()},
        period_expenses=period_expenses,
        expense_categories=Expense.CATEGORIES,
        top_products=top_products,
        sales_by_rep=sales_by_rep,
        payment_breakdown=payment_breakdown,
        chart_labels=chart_labels,
        chart_revenue=chart_revenue,
        current_year=current_year,
        yearly_labels=yearly_labels,
        yearly_rev=yearly_rev,
        yearly_cogs=yearly_cogs_l,
        yearly_exp=yearly_exp_l,
        yearly_gross=yearly_gross_l,
        yearly_net=yearly_net_l,
    )


# ─── Expenses ─────────────────────────────────────────────────────────────────

@admin_bp.route('/<slug>/admin/expenses/add', methods=['POST'])
@login_required
def add_expense(slug):
    biz = _load_biz(slug)
    period = request.form.get('period', 'month')
    category = request.form.get('category', '').strip()
    description = request.form.get('description', '').strip()
    amount_raw = request.form.get('amount', '').strip()
    date_raw = request.form.get('expense_date', '').strip()

    errors = []
    if not category:
        errors.append('Category is required.')
    try:
        amount = Decimal(amount_raw)
        if amount <= 0:
            raise ValueError
    except Exception:
        errors.append('Enter a valid positive amount.')
        amount = Decimal('0')
    try:
        exp_date = datetime.strptime(date_raw, '%Y-%m-%d').date() if date_raw else date_type.today()
    except ValueError:
        errors.append('Invalid date.')
        exp_date = date_type.today()

    if errors:
        for e in errors:
            flash(e, 'danger')
    else:
        db.session.add(Expense(
            business_id=biz.id,
            category=category,
            description=description,
            amount=amount,
            expense_date=exp_date,
            recorded_by=current_user.id,
        ))
        db.session.commit()
        flash(f'Expense recorded: {category} — {biz.currency_symbol}{amount:,.2f}', 'success')

    return redirect(url_for('admin.reports', slug=slug, period=period, t='fin'))


@admin_bp.route('/<slug>/admin/expenses/<int:expense_id>/delete', methods=['POST'])
@login_required
def delete_expense(slug, expense_id):
    biz = _load_biz(slug)
    exp = Expense.query.filter_by(id=expense_id, business_id=biz.id).first_or_404()
    period = request.form.get('period', 'month')
    db.session.delete(exp)
    db.session.commit()
    flash('Expense deleted.', 'success')
    return redirect(url_for('admin.reports', slug=slug, period=period, t='fin'))


# ─── Users ───────────────────────────────────────────────────────────────────

@admin_bp.route('/<slug>/admin/users')
@login_required
def users(slug):
    biz = _load_biz(slug)
    user_list = User.query.filter_by(business_id=biz.id).order_by(User.role, User.username).all()
    return render_template('admin/users.html', business=biz, users=user_list)


@admin_bp.route('/<slug>/admin/users/new', methods=['GET', 'POST'])
@login_required
def new_user(slug):
    biz = _load_biz(slug)
    sub = biz.subscription
    features = sub.features if sub else PLAN_LIMITS['free']
    max_users = features.get('max_users')
    if max_users is not None:
        current_count = User.query.filter_by(business_id=biz.id, is_active=True).count()
        if current_count >= max_users:
            flash(f'Your {features["label"]} plan allows up to {max_users} users. '
                  f'Upgrade to add more.', 'warning')
            return redirect(url_for('admin.upgrade', slug=slug))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', 'sales_rep').strip()
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        email = request.form.get('email', '').strip() or None

        errors = []
        if not username:
            errors.append('Username is required.')
        if not password or len(password) < 6:
            errors.append('Password must be at least 6 characters.')
        if role not in ('admin', 'store_keeper', 'sales_rep'):
            errors.append('Invalid role.')
        if User.query.filter_by(business_id=biz.id, username=username).first():
            errors.append(f'Username "{username}" already exists.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/user_form.html', business=biz, user=None)

        user = User(
            business_id=biz.id,
            username=username,
            email=email,
            role=role,
            first_name=first_name,
            last_name=last_name,
        )
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash(f'User "{username}" ({role}) created.', 'success')
        return redirect(url_for('admin.users', slug=slug))

    return render_template('admin/user_form.html', business=biz, user=None)


@admin_bp.route('/<slug>/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_user(slug, user_id):
    biz = _load_biz(slug)
    user = User.query.filter_by(id=user_id, business_id=biz.id).first_or_404()

    if request.method == 'POST':
        user.first_name = request.form.get('first_name', '').strip()
        user.last_name = request.form.get('last_name', '').strip()
        user.email = request.form.get('email', '').strip() or None
        new_password = request.form.get('password', '').strip()
        if new_password:
            if len(new_password) < 6:
                flash('Password must be at least 6 characters.', 'danger')
                return render_template('admin/user_form.html', business=biz, user=user)
            user.set_password(new_password)
        if user.id != current_user.id:
            user.role = request.form.get('role', user.role)
            user.is_active = bool(request.form.get('is_active'))
        db.session.commit()
        flash(f'User "{user.username}" updated.', 'success')
        return redirect(url_for('admin.users', slug=slug))

    return render_template('admin/user_form.html', business=biz, user=user)


# ─── Work Sessions (Admin view) ──────────────────────────────────────────────

@admin_bp.route('/<slug>/admin/sessions')
@login_required
def sessions(slug):
    biz = _load_biz(slug)
    period = request.args.get('period', 'today')
    start, end = get_date_range(period)
    sess_list = (WorkSession.query
                 .filter(WorkSession.business_id == biz.id,
                         WorkSession.clock_in >= start,
                         WorkSession.clock_in <= end)
                 .order_by(WorkSession.clock_in.desc())
                 .all())
    return render_template('admin/sessions.html', business=biz, sessions=sess_list, period=period)


# ─── Price History ───────────────────────────────────────────────────────────

@admin_bp.route('/<slug>/admin/price-history')
@login_required
def price_history(slug):
    biz = _load_biz(slug)
    sub = biz.subscription
    features = sub.features if sub else PLAN_LIMITS['free']
    if not features.get('has_price_history'):
        flash('Price history is not available on your current plan. Upgrade to unlock it.', 'warning')
        return redirect(url_for('admin.upgrade', slug=slug))

    period = request.args.get('period', '').strip()
    query = PriceHistory.query.join(Product).filter(Product.business_id == biz.id)
    if period in ('today', 'week', 'month'):
        start, end = get_date_range(period)
        query = query.filter(PriceHistory.changed_at >= start, PriceHistory.changed_at <= end)
    history = query.order_by(PriceHistory.changed_at.desc()).limit(100).all()
    return render_template('admin/price_history.html', business=biz, history=history, period=period)


# ─── Business Settings ────────────────────────────────────────────────────────

@admin_bp.route('/<slug>/admin/settings', methods=['GET', 'POST'])
@login_required
def settings(slug):
    biz = _load_biz(slug)
    valid_currencies = {code: symbol for code, symbol, _ in Config.CURRENCIES}

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        currency = request.form.get('currency', '').strip()
        print_mode = request.form.get('print_mode', '').strip()
        printer_name = request.form.get('printer_name', '').strip()
        paper_width_mm = request.form.get('paper_width_mm', '').strip()

        if not name:
            flash('Business name is required.', 'danger')
            return render_template('admin/settings.html', business=biz)
        if currency not in valid_currencies:
            flash('Please select a valid currency.', 'danger')
            return render_template('admin/settings.html', business=biz)
        if print_mode not in ('browser', 'qz'):
            flash('Please select a valid print mode.', 'danger')
            return render_template('admin/settings.html', business=biz)
        if paper_width_mm not in ('58', '80'):
            flash('Please select a valid paper width.', 'danger')
            return render_template('admin/settings.html', business=biz)
        if print_mode == 'qz' and not printer_name:
            flash('Enter a printer name (or use "Detect Printers") before enabling direct thermal printing.', 'danger')
            return render_template('admin/settings.html', business=biz)

        biz.name = name
        biz.phone = phone
        biz.address = address
        biz.currency = currency
        biz.currency_symbol = valid_currencies[currency]
        biz.print_mode = print_mode
        biz.printer_name = printer_name or None
        biz.paper_width_mm = int(paper_width_mm)
        db.session.commit()
        flash('Business settings updated.', 'success')
        return redirect(url_for('admin.settings', slug=slug))

    return render_template('admin/settings.html', business=biz)


# ─── Upgrade / Pricing ───────────────────────────────────────────────────────

@admin_bp.route('/<slug>/admin/upgrade')
@login_required
def upgrade(slug):
    biz = _load_biz(slug)
    sub = biz.subscription
    return render_template('admin/upgrade.html', business=biz,
                           sub=sub, PLAN_LIMITS=PLAN_LIMITS)


# ─── Billing (Paystack) ───────────────────────────────────────────────────────

@admin_bp.route('/<slug>/admin/billing/initialize', methods=['POST'])
@login_required
def billing_initialize(slug):
    """Initialize a Paystack transaction and redirect to hosted checkout."""
    biz   = _load_biz(slug)
    plan  = request.form.get('plan', '').strip()
    cycle = request.form.get('billing_cycle', 'monthly').strip()

    if plan not in ('pro', 'premium') or cycle not in ('monthly', 'annual'):
        flash('Invalid plan selection.', 'danger')
        return redirect(url_for('admin.upgrade', slug=slug))

    prices     = current_app.config['PLAN_PRICES']
    amount     = prices[plan][cycle]
    currency   = current_app.config.get('PAYSTACK_CURRENCY', 'USD')
    secret_key = current_app.config.get('PAYSTACK_SECRET_KEY', '')

    if not secret_key:
        flash('Payment is not configured yet. Please contact support.', 'warning')
        return redirect(url_for('admin.upgrade', slug=slug))

    callback = url_for('admin.billing_verify', slug=slug, _external=True)
    payload  = {
        'email':        biz.email,
        'amount':       amount,
        'currency':     currency,
        'callback_url': callback,
        'metadata': {
            'business_id':   biz.id,
            'plan':          plan,
            'billing_cycle': cycle,
            'slug':          slug,
        },
    }

    try:
        resp = http_requests.post(
            'https://api.paystack.co/transaction/initialize',
            json=payload,
            headers={'Authorization': f'Bearer {secret_key}'},
            timeout=15,
        )
        data = resp.json()
    except Exception as exc:
        current_app.logger.error(f'Paystack init error: {exc}')
        flash('Could not reach payment gateway. Please try again.', 'danger')
        return redirect(url_for('admin.upgrade', slug=slug))

    if not data.get('status'):
        msg = data.get('message', 'Unknown error')
        flash(f'Payment initialization failed: {msg}', 'danger')
        return redirect(url_for('admin.upgrade', slug=slug))

    return redirect(data['data']['authorization_url'])


@admin_bp.route('/<slug>/admin/billing/verify')
@login_required
def billing_verify(slug):
    """Paystack redirects here after payment with ?reference=..."""
    biz       = _load_biz(slug)
    reference = request.args.get('reference', '').strip()

    if not reference:
        flash('Payment reference is missing.', 'danger')
        return redirect(url_for('admin.upgrade', slug=slug))

    secret_key = current_app.config.get('PAYSTACK_SECRET_KEY', '')

    try:
        resp = http_requests.get(
            f'https://api.paystack.co/transaction/verify/{reference}',
            headers={'Authorization': f'Bearer {secret_key}'},
            timeout=15,
        )
        data = resp.json()
    except Exception as exc:
        current_app.logger.error(f'Paystack verify error: {exc}')
        flash('Could not verify payment. Please contact support.', 'danger')
        return redirect(url_for('admin.upgrade', slug=slug))

    if not data.get('status') or data['data'].get('status') != 'success':
        flash('Payment was not successful. Please try again.', 'danger')
        return redirect(url_for('admin.upgrade', slug=slug))

    tx   = data['data']
    meta = tx.get('metadata', {})
    plan  = meta.get('plan')
    cycle = meta.get('billing_cycle', 'monthly')

    if plan not in ('pro', 'premium'):
        flash('Unrecognised plan in payment. Please contact support.', 'danger')
        return redirect(url_for('admin.upgrade', slug=slug))

    # Activate subscription
    sub = biz.subscription
    if not sub:
        sub = Subscription(business_id=biz.id)
        db.session.add(sub)

    sub.plan                  = plan
    sub.billing_cycle         = cycle
    sub.status                = 'active'
    sub.period_start          = datetime.utcnow()
    sub.period_end            = (datetime.utcnow() + timedelta(days=365)
                                 if cycle == 'annual'
                                 else datetime.utcnow() + timedelta(days=30))
    sub.plan_r5d_sent         = False
    sub.plan_r2d_sent         = False
    sub.plan_r1d_sent         = False
    sub.paystack_reference    = reference
    customer = tx.get('customer', {})
    if customer.get('customer_code'):
        sub.paystack_customer_code = customer['customer_code']

    db.session.commit()

    flash(f'Payment successful! Your {plan.title()} ({cycle}) plan is now active.', 'success')
    return redirect(url_for('admin.billing_success', slug=slug,
                            plan=plan, cycle=cycle))


@admin_bp.route('/<slug>/admin/billing/success')
@login_required
def billing_success(slug):
    biz   = _load_biz(slug)
    plan  = request.args.get('plan', 'pro')
    cycle = request.args.get('cycle', 'monthly')
    return render_template('admin/billing_success.html',
                           business=biz, plan=plan, cycle=cycle,
                           sub=biz.subscription)
